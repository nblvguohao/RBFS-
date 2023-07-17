#-*- codeing=utf-8 -*-
#@Time:2022/3/2 10:27
#@Author:lvguohao
#@File:txtToExcel.py
#@software:PyCharm
#!/bin/env python
# -*- encoding: utf-8 -*-
#-------------------------------------------------------------------------------
# Purpose:     txt转换成Excel
# Author:      zhoujy
# Created:     2013-05-07
# update:      2013-05-07
#-------------------------------------------------------------------------------
import datetime
import time
import os
import sys
import xlwt #需要的模块
import openpyxl
import pandas as pd
import numpy as np
import xlsxwriter
from fuzzywuzzy import fuzz
import re

def txt2xls(filename,xlsname):  #文本转换成xls的函数，filename 表示一个要被转换的txt文本，xlsname 表示转换后的文件名
    print('converting xls ... ')
    f = open(filename)   #打开txt文本进行读取
    x = 1                #在excel开始写的位置（y）
    y = 1                #在excel开始写的位置（x）
    xls=openpyxl.Workbook()
    result = []
    sheet = xls.create_sheet(index=0) #生成excel的方法，声明excel
    while True:  #循环，读取文本里面的所有内容
        line = f.readline() #一行一行读取
        if not line:  #如果没有内容，则退出循环
            break
        for i in line.split('\t'):#读取出相应的内容写到x
            item=i.strip()
            # obj["protein"] = item
            # print(item)
            sheet.cell(x,y).value = item
            y += 1 #另起一列
        x += 1 #另起一行
        y = 1  #初始成第一列
    f.close()
    xls.save(xlsname+'.xls') #保存

if __name__ == "__main__":
    # filename = sys.argv[1]
    # xlsname  = sys.argv[2]
    #txt文件转成excel文件
    # txt2xls(r'C:\Users\123\Desktop\lvguohao\数据集\yeast_compartment_integrated_full.tsv',r'C:\Users\123\Desktop\lvguohao\data\dataset\Subcellular')

    # test = "Nuclear outer membrane-endoplasmic reticulum membrane network"
    # p = fuzz.partial_ratio("Endoplasmic", test)
    # match = re.search("Endoplasmic", test, re.IGNORECASE)
    # if match != None:
    #     print(match)
    # else:
    #     print("miss")


    path = os.path.join(r'C:\Users\123\Desktop\lvguohao\data\dataset\Subcellular.xls')
    data = pd.read_excel(path)
    List = (np.array(data)).tolist()
    result_arr = []
    for item in List:
        obj = {}
        if item[0].__len__() == 5 or item[0].__len__() == 7 or item[0].split("-").__len__()>1:
            obj["protein"] = item[0]
            obj["id"] = item[1]
            obj["Location"] = item[3]
            obj["probility"] = item[4]
            result_arr.append(obj)
    print(result_arr.__len__())

    subcellularLocalization = ['Endoplasmic', 'Cytoskeleton', 'Golgi', 'Cytosol', 'Vacuole', 'Mitochondrion',
                               'Endosome', 'Plasma', 'Nucleus', 'Peroxisome', 'Extracellular']  # 11个亚细胞定位区室

    trans = []

    for item in result_arr:
        obj = {}
        for item1 in subcellularLocalization:
            match = re.search(item1, item["Location"], re.IGNORECASE)
            if match != None:
                obj["protein"] = "'"+item["protein"]+"'"
                obj["position"] = "'"+item1+"'"
                trans.append(obj)

    print(trans.__len__())

    result_arr = trans

    workbook = xlsxwriter.Workbook('C:\\Users\\123\\Desktop\\lvguohao\\data\\dataset\\subcellularLocation_0302.xlsx')
    worksheet = workbook.add_worksheet()
    bold_format = workbook.add_format({'bold': True})
    worksheet.set_column(1, 1, 15)
    worksheet.write('A1', 'protein', bold_format)
    # worksheet.write('B1', 'id', bold_format)
    worksheet.write('B1', 'position', bold_format)
    # worksheet.write('D1', 'probility', bold_format)
    row = 1
    col = 0
    for item in (result_arr):
        worksheet.write_string(row, col, str(item['protein']))
        # worksheet.write_string(row, col + 1, str(item['id']))
        worksheet.write_string(row, col + 1, str(item['position']))
        # worksheet.write_string(row, col + 3, str(item['probility']))
        row += 1
    workbook.close()